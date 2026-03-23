from __future__ import annotations

from datetime import date as date_type
from decimal import Decimal
from typing import Optional

from PySide6.QtCore import Qt, Signal
from PySide6.QtGui import QColor, QFont, QKeySequence, QAction
from PySide6.QtWidgets import (
    QFileDialog,
    QFrame,
    QGridLayout,
    QGroupBox,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QMessageBox,
    QPushButton,
    QSplitter,
    QTableWidget,
    QTableWidgetItem,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)

from app.gui.table_helpers import style_table, show_empty_state
from app.services.customer_service import CustomerService
from app.services.order_service import OrderService


# Status boje za narudžbe
_ORDER_STATUS_COLOR = {
    "active":    "#059669",
    "completed": "#6b7280",
    "cancelled": "#dc2626",
}
_ORDER_STATUS_LABEL = {
    "active":    "Aktivna",
    "completed": "Završena",
    "cancelled": "Otkazana",
}


class CustomersPage(QWidget):
    """
    Stranica za upravljanje kupcima.

    Layout (QSplitter):
    - Lijevo: lista kupaca sa pretragom, dugmad Novi/Obriši
    - Desno: kartica kupca (podaci + forma za izmjenu) + tabela narudžbi
    """

    def __init__(self) -> None:
        super().__init__()
        self._selected_customer_id: Optional[int] = None

        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        splitter = QSplitter(Qt.Horizontal)
        splitter.setHandleWidth(1)
        splitter.setStyleSheet("QSplitter::handle { background: #e5e7eb; }")

        splitter.addWidget(self._build_left_panel())
        splitter.addWidget(self._build_right_panel())
        splitter.setSizes([340, 900])
        splitter.setStretchFactor(0, 0)
        splitter.setStretchFactor(1, 1)

        root.addWidget(splitter)

        # Ctrl+N = novi kupac (koristi QAction jer QShortcut nije dostupan)
        new_action = QAction(self)
        new_action.setShortcut(QKeySequence("Ctrl+N"))
        new_action.triggered.connect(self._new_customer)
        self.addAction(new_action)

        self.load_customers()

    # ------------------------------------------------------------------
    # Izgradnja UI — lijevi panel (lista kupaca)
    # ------------------------------------------------------------------

    def _build_left_panel(self) -> QFrame:
        panel = QFrame()
        panel.setProperty("card", True)
        panel.setMinimumWidth(280)
        panel.setMaximumWidth(380)

        layout = QVBoxLayout(panel)
        layout.setContentsMargins(14, 14, 14, 14)
        layout.setSpacing(10)

        # Naslov
        title = QLabel("Kupci")
        title.setProperty("sectionTitle", True)
        layout.addWidget(title)

        # Pretraga
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("🔍 Pretraži po imenu, tel, gradu...")
        self.search_input.textChanged.connect(self.load_customers)
        layout.addWidget(self.search_input)

        # Lista kupaca
        self.customers_table = QTableWidget(0, 2)
        self.customers_table.setHorizontalHeaderLabels(["Ime i prezime", "Grad"])
        hh = self.customers_table.horizontalHeader()
        hh.setSectionResizeMode(0, QHeaderView.Stretch)
        hh.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.customers_table.verticalHeader().setVisible(False)
        self.customers_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.customers_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.customers_table.setAlternatingRowColors(True)
        self.customers_table.itemSelectionChanged.connect(self._on_customer_selected)
        layout.addWidget(self.customers_table, 1)

        # Dugmad
        btn_row = QHBoxLayout()
        btn_row.setSpacing(8)

        self.new_btn = QPushButton("+ Novi kupac")
        self.new_btn.setProperty("primary", True)
        self.new_btn.setToolTip("Unesi novog kupca (Ctrl+N)")
        self.new_btn.clicked.connect(self._new_customer)

        self.delete_btn = QPushButton("Obriši")
        self.delete_btn.setProperty("secondary", True)
        self.delete_btn.setToolTip("Trajno obriši odabranog kupca")
        self.delete_btn.setEnabled(False)
        self.delete_btn.clicked.connect(self._delete_customer)

        btn_row.addWidget(self.new_btn, 1)
        btn_row.addWidget(self.delete_btn)
        layout.addLayout(btn_row)

        # Izvoz u Excel
        export_btn = QPushButton("📊 Izvezi u Excel")
        export_btn.setProperty("secondary", True)
        export_btn.setToolTip("Izvezi prikazane kupce u Excel fajl")
        export_btn.clicked.connect(self._export_to_excel)
        layout.addWidget(export_btn)

        # Brojač
        self.count_label = QLabel("")
        self.count_label.setStyleSheet("color: #9ca3af; font-size: 11px;")
        self.count_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.count_label)

        return panel

    # ------------------------------------------------------------------
    # Izgradnja UI — desni panel (kartica kupca)
    # ------------------------------------------------------------------

    def _build_right_panel(self) -> QFrame:
        panel = QFrame()
        panel.setProperty("card", True)

        layout = QVBoxLayout(panel)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(16)

        # Placeholder kad nema odabranog kupca
        self.placeholder_label = QLabel("← Odaberi kupca iz liste ili klikni \"+ Novi kupac\"")
        self.placeholder_label.setStyleSheet("color: #9ca3af; font-size: 14px;")
        self.placeholder_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.placeholder_label)

        # Kartica kupca (skrivena dok nema selekcije)
        self.customer_card = QWidget()
        self.customer_card.hide()
        card_layout = QVBoxLayout(self.customer_card)
        card_layout.setContentsMargins(0, 0, 0, 0)
        card_layout.setSpacing(16)

        # --- Forma za podatke kupca ---
        form_group = QGroupBox("Podaci kupca")
        form_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold; color: #374151;
                border: 1px solid #e5e7eb; border-radius: 8px;
                margin-top: 8px; padding-top: 12px;
            }
            QGroupBox::title {
                subcontrol-origin: margin; left: 12px; padding: 0 6px;
            }
        """)
        form_layout_outer = QVBoxLayout(form_group)

        grid = QGridLayout()
        grid.setHorizontalSpacing(12)
        grid.setVerticalSpacing(10)
        grid.setColumnStretch(1, 1)
        grid.setColumnStretch(3, 1)

        self.full_name_input = QLineEdit()
        self.phone_input = QLineEdit()
        self.city_input = QLineEdit()
        self.address_input = QLineEdit()
        self.note_input = QTextEdit()
        self.note_input.setMaximumHeight(70)
        self.note_input.setPlaceholderText("Opciona napomena...")

        # 2 kolone
        grid.addWidget(QLabel("Ime i prezime *"), 0, 0)
        grid.addWidget(self.full_name_input, 0, 1)
        grid.addWidget(QLabel("Telefon"), 0, 2)
        grid.addWidget(self.phone_input, 0, 3)

        grid.addWidget(QLabel("Mjesto"), 1, 0)
        grid.addWidget(self.city_input, 1, 1)
        grid.addWidget(QLabel("Adresa"), 1, 2)
        grid.addWidget(self.address_input, 1, 3)

        grid.addWidget(QLabel("Napomena"), 2, 0)
        grid.addWidget(self.note_input, 2, 1, 1, 3)

        form_layout_outer.addLayout(grid)

        # Dugmad za formu
        form_btns = QHBoxLayout()
        self.save_btn = QPushButton("💾  Sačuvaj izmjene")
        self.save_btn.setProperty("primary", True)
        self.save_btn.clicked.connect(self._save_customer)

        self.cancel_btn = QPushButton("Odustani")
        self.cancel_btn.setProperty("secondary", True)
        self.cancel_btn.clicked.connect(self._cancel_edit)
        self.cancel_btn.hide()  # Vidljiv samo pri novom unosu

        form_btns.addWidget(self.save_btn)
        form_btns.addWidget(self.cancel_btn)
        form_btns.addStretch(1)
        form_layout_outer.addLayout(form_btns)

        card_layout.addWidget(form_group)

        # --- Tabela narudžbi ---
        orders_group = QGroupBox("Narudžbe ovog kupca")
        orders_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold; color: #374151;
                border: 1px solid #e5e7eb; border-radius: 8px;
                margin-top: 8px; padding-top: 12px;
            }
            QGroupBox::title {
                subcontrol-origin: margin; left: 12px; padding: 0 6px;
            }
        """)
        orders_layout = QVBoxLayout(orders_group)

        self.orders_table = QTableWidget(0, 8)
        self.orders_table.setHorizontalHeaderLabels([
            "Br. ugovora", "Artikal", "Cijena (EUR)",
            "Plaćeno", "Preostalo (EUR)", "Završava", "Kasni", "Status"
        ])
        oh = self.orders_table.horizontalHeader()
        oh.setSectionResizeMode(0, QHeaderView.ResizeToContents)  # Br. ugovora
        oh.setSectionResizeMode(1, QHeaderView.Stretch)            # Artikal
        oh.setSectionResizeMode(2, QHeaderView.ResizeToContents)  # Cijena
        oh.setSectionResizeMode(3, QHeaderView.ResizeToContents)  # Plaćeno
        oh.setSectionResizeMode(4, QHeaderView.ResizeToContents)  # Preostalo
        oh.setSectionResizeMode(5, QHeaderView.ResizeToContents)  # Završava
        oh.setSectionResizeMode(6, QHeaderView.ResizeToContents)  # Kasni
        oh.setSectionResizeMode(7, QHeaderView.ResizeToContents)  # Status
        self.orders_table.verticalHeader().setVisible(False)
        self.orders_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.orders_table.setAlternatingRowColors(True)
        orders_layout.addWidget(self.orders_table)

        card_layout.addWidget(orders_group, 1)
        layout.addWidget(self.customer_card, 1)

        return panel

    # ------------------------------------------------------------------
    # Učitavanje podataka
    # ------------------------------------------------------------------

    def load_customers(self) -> None:
        customers = CustomerService.list_customers(self.search_input.text())

        self.count_label.setText(f"{len(customers)} kupaca")

        if not customers:
            show_empty_state(self.customers_table, "Nema kupaca")
            return

        self.customers_table.setRowCount(0)
        self.customers_table.setRowCount(len(customers))

        for row, customer in enumerate(customers):
            name_item = QTableWidgetItem(customer.full_name)
            name_item.setData(Qt.UserRole, customer.id)
            city_item = QTableWidgetItem(customer.city or "")

            self.customers_table.setItem(row, 0, name_item)
            self.customers_table.setItem(row, 1, city_item)

            # Vrati selekciju ako je isti kupac bio odabran
            if customer.id == self._selected_customer_id:
                self.customers_table.selectRow(row)

    def _load_orders_for_customer(self, customer_id: int) -> None:
        """Učitava narudžbe za kupca koristeći DTO objekte."""
        orders = OrderService.get_orders_for_customer(customer_id)

        if not orders:
            show_empty_state(self.orders_table, "Ovaj kupac nema narudžbi")
            return

        self.orders_table.setRowCount(0)
        self.orders_table.setRowCount(len(orders))

        for row, order in enumerate(orders):
            status_raw = (
                order.status.value if hasattr(order.status, "value") else str(order.status)
            )
            status_bs = {
                "active":    "Aktivna",
                "completed": "Završena",
                "cancelled": "Otkazana",
            }.get(status_raw, status_raw)

            # --- Izračuni iz DTO-a (već izračunato u service-u) ---
            installments = order.installments or []
            total_inst   = len(installments)

            # Broj plaćenih rata (status == paid)
            paid_count = sum(
                1 for i in installments
                if (i.status.value if hasattr(i.status, "value") else str(i.status)) == "paid"
            )

            # Ukupno uplaćeno EUR - koristi paid_amount iz DTO-a
            total_paid = sum(i.paid_amount for i in installments)
            preostalo = max(Decimal("0.00"), order.total_price_snapshot - total_paid)

            # Datum posljednje rate (završetak otplate)
            last_due = None
            if installments:
                last_due = max((i.due_date for i in installments), default=None)

            # Kasni: postoji li rata koja je overdue
            kasni = any(
                (i.status.value if hasattr(i.status, "value") else str(i.status)) == "overdue"
                for i in installments
            )

            # --- Broj ugovora ---
            contract = order.contract_number or f"#{order.id}"

            # --- Popunjavanje ćelija ---
            row_data = [
                (contract,                                              Qt.AlignCenter),
                (order.product_name_snapshot,                           Qt.AlignLeft),
                (f"{order.total_price_snapshot:.2f} EUR",               Qt.AlignRight),
                (f"{paid_count} / {total_inst}",                        Qt.AlignCenter),
                (f"{preostalo:.2f} EUR",                               Qt.AlignRight),
                (last_due.strftime("%d.%m.%Y") if last_due else "—",  Qt.AlignCenter),
                ("⚠" if kasni else "—",                               Qt.AlignCenter),
                (status_bs,                                             Qt.AlignCenter),
            ]

            for col, (val, align) in enumerate(row_data):
                item = QTableWidgetItem(str(val))
                item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
                item.setTextAlignment(align | Qt.AlignVCenter)

                # Br. ugovora — plava ako je pravi broj
                if col == 0 and order.contract_number:
                    item.setForeground(QColor("#1d4ed8"))
                    font = QFont()
                    font.setBold(True)
                    item.setFont(font)

                # Plaćeno — zelena kad je sve plaćeno
                if col == 3 and paid_count == total_inst and total_inst > 0:
                    item.setForeground(QColor("#059669"))
                    font = QFont()
                    font.setBold(True)
                    item.setFont(font)

                # Preostalo — zelena ako 0, narandžasta ako ima duga
                if col == 4:
                    if preostalo == Decimal("0.00"):
                        item.setForeground(QColor("#059669"))
                    else:
                        item.setForeground(QColor("#d97706"))

                # Kasni — crvena ⚠
                if col == 6 and kasni:
                    item.setForeground(QColor("#dc2626"))
                    font = QFont()
                    font.setBold(True)
                    item.setFont(font)

                # Status boja
                if col == 7:
                    color = {
                        "Aktivna":  "#059669",
                        "Završena": "#6b7280",
                        "Otkazana": "#dc2626",
                    }.get(status_bs, "#374151")
                    item.setForeground(QColor(color))

                self.orders_table.setItem(row, col, item)

        self.orders_table.resizeColumnsToContents()

    # ------------------------------------------------------------------
    # Akcije
    # ------------------------------------------------------------------

    def _on_customer_selected(self) -> None:
        row = self.customers_table.currentRow()
        if row < 0:
            return
        item = self.customers_table.item(row, 0)
        if not item:
            return

        customer_id = item.data(Qt.UserRole)
        customer = CustomerService.get_customer(customer_id)
        if not customer:
            return

        self._selected_customer_id = customer_id
        self.delete_btn.setEnabled(True)

        # Popuni formu
        self.full_name_input.setText(customer.full_name)
        self.phone_input.setText(customer.phone or "")
        self.city_input.setText(customer.city or "")
        self.address_input.setText(customer.address or "")
        self.note_input.setPlainText(customer.note or "")

        # Prikaži karticu
        self.placeholder_label.hide()
        self.customer_card.show()
        self.cancel_btn.hide()

        # Učitaj narudžbe
        self._load_orders_for_customer(customer_id)

    def _new_customer(self) -> None:
        """Pripremi formu za unos novog kupca."""
        self._selected_customer_id = None
        self.customers_table.clearSelection()
        self.delete_btn.setEnabled(False)

        self.full_name_input.clear()
        self.phone_input.clear()
        self.city_input.clear()
        self.address_input.clear()
        self.note_input.clear()
        self.orders_table.setRowCount(0)

        self.placeholder_label.hide()
        self.customer_card.show()
        self.cancel_btn.show()
        self.full_name_input.setFocus()

        # Naslov grupe da označi da je novi unos
        show_empty_state(self.orders_table, "Novi kupac — još nema narudžbi")

    def _save_customer(self) -> None:
        try:
            payload = dict(
                full_name=self.full_name_input.text(),
                phone=self.phone_input.text(),
                city=self.city_input.text(),
                address=self.address_input.text(),
                note=self.note_input.toPlainText(),
            )
            if self._selected_customer_id is None:
                customer = CustomerService.create_customer(**payload)
                self._selected_customer_id = customer.id
            else:
                CustomerService.update_customer(self._selected_customer_id, **payload)
        except Exception as exc:
            QMessageBox.warning(self, "Greška", str(exc))
            return

        self.cancel_btn.hide()
        self.load_customers()

    def _cancel_edit(self) -> None:
        """Odustani od novog unosa."""
        self._selected_customer_id = None
        self.placeholder_label.show()
        self.customer_card.hide()
        self.customers_table.clearSelection()
        self.delete_btn.setEnabled(False)

    def _delete_customer(self) -> None:
        if self._selected_customer_id is None:
            return

        name = self.full_name_input.text() or "odabranog kupca"
        reply = QMessageBox.question(
            self,
            "Potvrda brisanja",
            f"Da li sigurno želiš obrisati kupca '{name}'?\n\n"
            "Kupac se može obrisati samo ako nema narudžbi.\n"
            "Ova radnja se ne može poništiti.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if reply != QMessageBox.Yes:
            return

        try:
            CustomerService.delete_customer(self._selected_customer_id)
        except Exception as exc:
            QMessageBox.warning(self, "Greška", str(exc))
            return

        self._selected_customer_id = None
        self.placeholder_label.show()
        self.customer_card.hide()
        self.delete_btn.setEnabled(False)
        self.load_customers()

    def _export_to_excel(self) -> None:
        """Izvezi trenutno prikazane kupce u Excel fajl."""
        customers = CustomerService.list_customers(self.search_input.text())
        if not customers:
            QMessageBox.information(self, "Izvoz", "Nema kupaca za izvoz.")
            return

        path, _ = QFileDialog.getSaveFileName(
            self, "Sačuvaj Excel fajl", "kupci.xlsx", "Excel fajlovi (*.xlsx)"
        )
        if not path:
            return

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Kupci"

            # Stilovi
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill("solid", fgColor="1F3864")
            header_align = Alignment(horizontal="center", vertical="center")
            thin = Side(style="thin", color="D1D5DB")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            alt_fill = PatternFill("solid", fgColor="F9FAFB")

            # Zaglavlje
            headers = ["Ime i prezime", "Telefon", "Grad", "Adresa", "Napomena"]
            col_widths = [30, 18, 20, 30, 40]
            for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = border
                ws.column_dimensions[cell.column_letter].width = w
            ws.row_dimensions[1].height = 24

            # Podaci
            for row_idx, c in enumerate(customers, start=2):
                row_fill = PatternFill("solid", fgColor="FFFFFF") if row_idx % 2 == 0 else alt_fill
                values = [
                    c.full_name,
                    c.phone or "",
                    c.city or "",
                    c.address or "",
                    c.note or "",
                ]
                for col, val in enumerate(values, start=1):
                    cell = ws.cell(row=row_idx, column=col, value=val)
                    cell.fill = row_fill
                    cell.border = border
                    cell.alignment = Alignment(vertical="center", wrap_text=(col == 5))
                ws.row_dimensions[row_idx].height = 18

            # Freeze zaglavlje
            ws.freeze_panes = "A2"

            wb.save(path)
            QMessageBox.information(
                self, "Izvoz uspješan",
                f"Izvezeno {len(customers)} kupaca:\n{path}"
            )
        except Exception as e:
            QMessageBox.critical(self, "Greška pri izvozu", str(e))

    def on_activate(self) -> None:
        self.load_customers()
