from __future__ import annotations

from datetime import date
from decimal import Decimal
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.database.database import session_scope
from app.database.models import Campaign, Installment, Order, Payment
from sqlalchemy import select
from sqlalchemy.orm import joinedload


# ---------------------------------------------------------------------------
# Bosanski nazivi mjeseci (kratica) — ne oslanjamo se na locale
# ---------------------------------------------------------------------------
_BS_MONTHS = {
    1: "jan", 2: "feb", 3: "mar", 4: "apr",
    5: "maj", 6: "jun", 7: "jul", 8: "aug",
    9: "sep", 10: "okt", 11: "nov", 12: "dec",
}


def _fmt_date(d: date) -> str:
    """Formatira datum u bosanski kratki format: '28.feb.'"""
    return f"{d.day}.{_BS_MONTHS[d.month]}."


def _fmt_num(val: Decimal) -> str:
    """Formatira broj u bosanski format: 1.234,56"""
    return f"{val:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


# ---------------------------------------------------------------------------
# Stilovi
# ---------------------------------------------------------------------------

def _border() -> Border:
    t = Side(style="thin")
    return Border(left=t, right=t, top=t, bottom=t)


def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _font(bold: bool = False, size: int = 9,
          color: str = "000000", name: str = "Calibri") -> Font:
    return Font(bold=bold, size=size, color=color, name=name)


def _align(h: str = "center", v: str = "center",
           wrap: bool = True) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# Gotove kombinacije
FILL_HEADER  = _fill("1F3864")   # tamno plava — zaglavlje kolona
FILL_DATES   = _fill("BDD7EE")   # svjetlo plava — red s datumima
FILL_UKUPNO  = _fill("D9E1F2")   # još svjetlija — UKUPNO red
FILL_ALT_ROW = _fill("F5F5F5")   # lagano siva — parni redovi


# ---------------------------------------------------------------------------
# Dohvat podataka iz baze
# ---------------------------------------------------------------------------

def _get_report_data(campaign_id: int) -> dict:
    """
    Vraća strukturirani dict sa svim narudžbama kampanje,
    ratama i uplatama — spreman za pisanje u Excel.
    """
    with session_scope() as session:
        campaign = session.get(Campaign, campaign_id)
        if not campaign:
            raise ValueError(f"Kampanja ID={campaign_id} ne postoji.")

        stmt = (
            select(Order)
            .where(Order.campaign_id == campaign_id)
            .options(
                joinedload(Order.customer),
                joinedload(Order.installments).joinedload(Installment.payments),
            )
            .order_by(Order.id.asc())
        )
        orders = list(session.execute(stmt).scalars().unique())

        rows = []
        for idx, order in enumerate(orders, start=1):
            installments = sorted(
                order.installments, key=lambda i: i.installment_number
            )
            inst_data = []
            total_paid = Decimal("0.00")

            for inst in installments:
                paid = sum(
                    (p.amount for p in inst.payments), Decimal("0.00")
                )
                total_paid += paid
                inst_data.append({
                    "number":   inst.installment_number,
                    "due_date": inst.due_date,
                    "amount":   inst.amount,
                    "paid":     paid,
                })

            total_amount = order.total_price_snapshot
            remaining = total_amount - total_paid

            rows.append({
                "rbr":          idx,
                # Koristimo contract_number (broj ugovora iz originala),
                # fallback na order.id ako ga nema
                "br_ugovora":   order.contract_number or str(order.id),
                "customer":     order.customer.full_name if order.customer else "?",
                "product":      order.product_name_snapshot,
                "total_amount": total_amount,
                "inst_count":   order.installments_count,
                "installments": inst_data,
                "total_paid":   total_paid,
                "remaining":    max(remaining, Decimal("0.00")),
            })

        max_inst = max((len(r["installments"]) for r in rows), default=1)

        return {
            "campaign_name": campaign.name,
            "rows":          rows,
            "max_inst":      max_inst,
        }


# ---------------------------------------------------------------------------
# Generisanje Excel fajla
# ---------------------------------------------------------------------------

def generate_naplata_excel(
    campaign_id: int,
    output_path: Path,
    agent_code: str = "",
    agent_name: str = "KRUNIĆ STOJANOVIĆ SANJA",
) -> Path:
    """
    Generira Excel izvještaj naplate koji odgovara originalnom
    formatu 'EVIDENCIJA O UPLATAMA RATA'.

    Kolumni raspored (identičan originalu):
      A  — prazan razmak
      B  — R.br.
      C  — Br. ugovora i datum
      D  — Prezime i ime
      E  — Šifra proizvoda
      F  — Vrijed. (EUR)
      G  — Br. rata
      H … H+max_inst-1 — rate I, II, III...
      H+max_inst       — Ukupno (EUR)
      H+max_inst+1     — Preostalo (EUR)
    """
    data      = _get_report_data(campaign_id)
    rows      = data["rows"]
    max_inst  = data["max_inst"]
    camp_name = data["campaign_name"]

    # ----------------------------------------------------------------
    # Indeksi kolona (1-based, openpyxl standard)
    # ----------------------------------------------------------------
    COL_RBR       = 2
    COL_UGOVOR    = 3
    COL_KUPAC     = 4
    COL_PROIZVOD  = 5
    COL_VRIJED    = 6
    COL_BR_RATA   = 7
    COL_INST_1    = 8
    COL_UKUPNO    = COL_INST_1 + max_inst
    COL_PREOSTALO = COL_UKUPNO + 1
    LAST_COL      = COL_PREOSTALO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Evidencija naplate"

    # ----------------------------------------------------------------
    # Širine kolona
    # ----------------------------------------------------------------
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions[get_column_letter(COL_RBR)].width      = 5
    ws.column_dimensions[get_column_letter(COL_UGOVOR)].width   = 13
    ws.column_dimensions[get_column_letter(COL_KUPAC)].width    = 22
    ws.column_dimensions[get_column_letter(COL_PROIZVOD)].width = 26
    ws.column_dimensions[get_column_letter(COL_VRIJED)].width   = 11
    ws.column_dimensions[get_column_letter(COL_BR_RATA)].width  = 6
    for i in range(max_inst):
        ws.column_dimensions[get_column_letter(COL_INST_1 + i)].width = 9
    ws.column_dimensions[get_column_letter(COL_UKUPNO)].width    = 11
    ws.column_dimensions[get_column_letter(COL_PREOSTALO)].width = 11

    # ================================================================
    # RED 1 — Naslov kompanije
    # ================================================================
    r = 1
    ws.row_dimensions[r].height = 20
    c = ws.cell(row=r, column=COL_RBR,
                value='EVIDENCIJA O UPLATAMA RATA ZA „MORE FOR LESS" d.o.o. '
                      'Istočno Sarajevo, Bosna i Hercegovina')
    c.font      = _font(bold=True, size=11)
    c.alignment = _align(h="left", wrap=False)
    ws.merge_cells(start_row=r, start_column=COL_RBR,
                   end_row=r, end_column=LAST_COL)

    # ================================================================
    # RED 2 — Info o saradniku
    # ================================================================
    r = 2
    ws.row_dimensions[r].height = 15

    ws.cell(row=r, column=COL_RBR,
            value="Saradnik:").font = _font(bold=True)
    ws.cell(row=r, column=COL_UGOVOR,
            value=agent_code).font = _font()
    ws.cell(row=r, column=COL_KUPAC,
            value=agent_name).font = _font(bold=True)
    ws.cell(row=r, column=COL_VRIJED,
            value="Kampanja:").font = _font(bold=True)
    ws.cell(row=r, column=COL_BR_RATA,
            value=camp_name).font = _font()

    # ================================================================
    # RED 3 — Zaglavlje kolona
    # ================================================================
    r = 3
    ws.row_dimensions[r].height = 30

    ROMAN = ["I", "II", "III", "IV", "V", "VI",
             "VII", "VIII", "IX", "X", "XI", "XII"]

    header_cells = {
        COL_RBR:       "R.\nbr.",
        COL_UGOVOR:    "Br. ugovora\ni datum",
        COL_KUPAC:     "Prezime i\nime",
        COL_PROIZVOD:  "Šifra\nproizvoda",
        COL_VRIJED:    "Vrijed.\n(EUR)",
        COL_BR_RATA:   "Br.\nrata",
        COL_UKUPNO:    "Ukupno\n(EUR)",
        COL_PREOSTALO: "Preostalo\n(EUR)",
    }
    for i in range(max_inst):
        header_cells[COL_INST_1 + i] = ROMAN[i] if i < len(ROMAN) else str(i + 1)

    for col, val in header_cells.items():
        c = ws.cell(row=r, column=col, value=val)
        c.font      = _font(bold=True, color="FFFFFF")
        c.fill      = FILL_HEADER
        c.alignment = _align()
        c.border    = _border()

    # ================================================================
    # RED 4 — Datumi dospijeća rata
    # ================================================================
    r = 4
    ws.row_dimensions[r].height = 16

    for col in (COL_RBR, COL_UGOVOR, COL_KUPAC, COL_PROIZVOD,
                COL_VRIJED, COL_BR_RATA, COL_UKUPNO, COL_PREOSTALO):
        c = ws.cell(row=r, column=col, value="")
        c.fill   = FILL_DATES
        c.border = _border()

    inst_dates: list[Optional[date]] = [None] * max_inst
    for row_data in rows:
        for inst in row_data["installments"]:
            idx = inst["number"] - 1
            if 0 <= idx < max_inst and inst_dates[idx] is None:
                inst_dates[idx] = inst["due_date"]

    for i, due in enumerate(inst_dates):
        label = _fmt_date(due) if due else ""
        c = ws.cell(row=r, column=COL_INST_1 + i, value=label)
        c.font      = _font(bold=True, size=8)
        c.fill      = FILL_DATES
        c.alignment = _align()
        c.border    = _border()

    # ================================================================
    # REDOVI PODATAKA
    # ================================================================
    for row_data in rows:
        r += 1
        ws.row_dimensions[r].height = 15
        is_even = row_data["rbr"] % 2 == 0

        def _cell(col: int, val, h: str = "center") -> openpyxl.cell.Cell:
            c = ws.cell(row=r, column=col, value=val)
            c.alignment = _align(h=h, wrap=False)
            c.border    = _border()
            c.font      = _font()
            if is_even:
                c.fill = FILL_ALT_ROW
            return c

        _cell(COL_RBR,      row_data["rbr"])
        _cell(COL_UGOVOR,   row_data["br_ugovora"])   # contract_number!
        _cell(COL_KUPAC,    row_data["customer"],    h="left")
        _cell(COL_PROIZVOD, row_data["product"],     h="left")
        _cell(COL_VRIJED,   _fmt_num(row_data["total_amount"]),  h="right")
        _cell(COL_BR_RATA,  row_data["inst_count"])

        # Prazne ćelije za neiskorištene kolone rata
        for i in range(max_inst):
            col = COL_INST_1 + i
            if ws.cell(row=r, column=col).value is None:
                c = ws.cell(row=r, column=col, value="")
                c.border = _border()
                c.font   = _font()
                if is_even:
                    c.fill = FILL_ALT_ROW

        # Uplaćene rate — plava boja za uplaćene iznose
        for inst in row_data["installments"]:
            col = COL_INST_1 + inst["number"] - 1
            if col >= COL_UKUPNO:
                continue
            if inst["paid"] > 0:
                c = ws.cell(row=r, column=col,
                            value=_fmt_num(inst["paid"]))
                c.alignment = _align(h="right", wrap=False)
                c.border    = _border()
                c.font      = _font(bold=True, color="0070C0")
                if is_even:
                    c.fill = FILL_ALT_ROW

        _cell(COL_UKUPNO,    _fmt_num(row_data["total_paid"]),  h="right")
        _cell(COL_PREOSTALO, _fmt_num(row_data["remaining"]),   h="right")

    # ================================================================
    # UKUPNO red
    # ================================================================
    r += 1
    ws.row_dimensions[r].height = 17

    ws.cell(row=r, column=COL_RBR, value="UKUPNO")
    ws.merge_cells(start_row=r, start_column=COL_RBR,
                   end_row=r, end_column=COL_BR_RATA)
    c = ws.cell(row=r, column=COL_RBR)
    c.font      = _font(bold=True, size=10)
    c.alignment = _align()
    c.fill      = FILL_UKUPNO
    c.border    = _border()

    for i in range(max_inst):
        col = COL_INST_1 + i
        col_total = sum(
            row_data["installments"][i]["paid"]
            if i < len(row_data["installments"]) else Decimal("0")
            for row_data in rows
        )
        val = _fmt_num(col_total) if col_total > 0 else ""
        c = ws.cell(row=r, column=col, value=val)
        c.font      = _font(bold=True)
        c.alignment = _align(h="right", wrap=False)
        c.fill      = FILL_UKUPNO
        c.border    = _border()

    grand_paid      = sum(rd["total_paid"]  for rd in rows)
    grand_remaining = sum(rd["remaining"]   for rd in rows)

    for col, val in ((COL_UKUPNO, grand_paid), (COL_PREOSTALO, grand_remaining)):
        c = ws.cell(row=r, column=col, value=_fmt_num(val))
        c.font      = _font(bold=True)
        c.alignment = _align(h="right", wrap=False)
        c.fill      = FILL_UKUPNO
        c.border    = _border()

    # ================================================================
    # Finalne postavke
    # ================================================================
    ws.freeze_panes = f"{get_column_letter(COL_KUPAC)}5"

    ws.page_setup.orientation  = "landscape"
    ws.page_setup.fitToPage    = True
    ws.page_setup.fitToWidth   = 1
    ws.page_setup.fitToHeight  = 0
    ws.print_title_rows        = "3:4"

    wb.save(output_path)
    return output_path
